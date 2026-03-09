"""
Trump Intel — Weekly Excel Report Generator
Gera data/report_latest.xlsx toda segunda-feira via GitHub Actions
5 abas: Dashboard, Watchlist, Teses, Gatilhos, Histórico
"""

import os
import json
import requests
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

FINNHUB_KEY = os.environ.get("FINNHUB_API_KEY", "")

# ── Paleta dark ─────────────────────────────────────────────────────────────
BG_DARK    = "0A0A0F"
BG_CARD    = "12121A"
AMBER      = "FFB800"
GREEN      = "00CC66"
RED        = "FF3366"
BLUE       = "0088CC"
WHITE      = "E8E8F0"
MUTED      = "5A5A7A"
BORDER_CLR = "1E1E2E"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(color=WHITE, bold=False, size=10, name="Arial"):
    return Font(color=color, bold=bold, size=size, name=name)

def border():
    s = Side(border_style="thin", color=BORDER_CLR)
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

# ── Finnhub ──────────────────────────────────────────────────────────────────
def get_quote(symbol: str) -> dict:
    if not FINNHUB_KEY:
        return {}
    try:
        r = requests.get(
            "https://finnhub.io/api/v1/quote",
            params={"symbol": symbol, "token": FINNHUB_KEY},
            timeout=5,
        )
        return r.json()
    except Exception:
        return {}

# ── Watchlist data ────────────────────────────────────────────────────────────
WATCHLIST = [
    # sym, finnhub_sym, name, rating, conviction, target, catalyst, sector, thesis
    ("GC=F",    "OANDA:XAU_USD", "Ouro Futuros",         "FORTE COMPRA", 95, 5000,  "Q4/2026",   "Commodities", "JP Morgan $5.000/oz. Fed dovish + BCs 900t/ano + déficit $2tri+."),
    ("SI=F",    "OANDA:XAG_USD", "Prata Futuros",         "FORTE COMPRA", 90, 58,    "Summit Xi", "Commodities", "China elevou a estratégico. Déficit estrutural. Solar +143%."),
    ("MP",      "MP",            "MP Materials",          "FORTE COMPRA", 88, 45,    "Summit Xi", "Defesa",      "Único produtor rare earth EUA em escala. DoD acionista."),
    ("REMX",    "REMX",          "VanEck Rare Earth ETF", "FORTE COMPRA", 82, 75,    "Nov/2026",  "Defesa",      "ETF exposição direta prazo nov/2026. Assimétrico."),
    ("NG=F",    "USOIL",         "Gás Natural",           "COMPRA",       85, 5.5,   "LNG deals", "Energia",     "Deal $33bi Japão. LNG exports. Trump incentiva."),
    ("LMT",     "LMT",           "Lockheed Martin",       "COMPRA",       83, 620,   "OTAN 5%",   "Defesa",      "OTAN 5% PIB. 4 frentes: Greenland, Venezuela, Irã, Ucrânia."),
    ("RTX",     "RTX",           "RTX Corp",              "COMPRA",       83, 140,   "OTAN 5%",   "Defesa",      "Rearmamento global. Catalisadores independentes de paz."),
    ("FCX",     "FCX",           "Freeport-McMoRan",      "COMPRA",       80, 55,    "S232 auto", "Commodities", "Maior produtor Cu EUA. Section 232 proteção doméstica."),
    ("NOC",     "NOC",           "Northrop Grumman",      "COMPRA",       80, 620,   "OTAN 5%",   "Defesa",      "Defesa espacial + GBSD. OTAN + Indo-Pacífico."),
    ("SLB",     "SLB",           "SLB Schlumberger",      "SELETIVO",     75, 48,    "Venezuela", "Energia",     "Play Venezuela serviços, não preço spot."),
    ("HAL",     "HAL",           "Halliburton",           "SELETIVO",     72, 35,    "Venezuela", "Energia",     "Venezuela caos pós-Maduro. HAL menos exposto que produtoras."),
    ("NVDA",    "NVDA",          "NVIDIA",                "SELETIVO",     75, 160,   "S232 semi", "Tech",        "CapEx IA $3tri. Semicondutores = Section 232."),
    ("CL=F",    "USOIL",         "Petróleo WTI",          "SELETIVO",     60, 80,    "Venezuela", "Energia",     "JP Morgan bearish. Venezuela ramp-up = oversupply potencial."),
    ("BTC-USD", "BINANCE:BTCUSDT","Bitcoin",              "COMPRA",       78, 150000,"Nov/2026",  "Crypto",      "Halving + ETFs spot + Trump pró-cripto. Ciclo Q4/2025–Q1/2026."),
    ("ETH-USD", "BINANCE:ETHUSDT","Ethereum",             "SELETIVO",     62, 6000,  "ETF staking","Crypto",     "ETF aprovado. SEC Atkins pró-cripto. Staking yield pendente."),
]

CATALYSTS = [
    ("31/Mar–2/Abr/2026", "Summit Trump–Xi",           "CRÍTICO", "Minerais raros na mesa. MP/REMX ±30% em 24h.",               "MP, REMX, SLV"),
    ("Abr/2026",          "Section 232 — Automotivo",  "ALTO",    "Peças automotivas incluídas em tarifas de aço/alumínio.",     "FCX, CLF, NUE"),
    ("Jun/2026",          "Tarifas Greenland",          "ALTO",    "Trump 25% países europeus que bloqueiem acesso.",             "LMT, RTX"),
    ("Nov/2026",          "TRIPLE EVENT",               "CRÍTICO", "Trégua EUA-China + pausa minerais raros + Midterms. 3 em 1.", "MP, REMX, GLD"),
    ("2027",              "Revisão USMCA",              "MÉDIO",   "México e Canadá vulneráveis a pressões tarifárias.",          "FCX, MPC"),
    ("Q4/2026",           "Ouro $5.000/oz",             "ALTO",    "Fed dovish + BCs 900t/ano + déficit $2tri+.",                 "GLD, GC=F, SLV"),
]

TESES = [
    ("Venezuela",       "Doutrina Monroe 2.0", "Maduro capturado jan/2026. EUA corre Venezuela. 17% reservas mundiais. Condição: cortar China/Rússia/Irã.",   "SLB, HAL — SERVIÇOS (não preço spot)"),
    ("China — Raros",   "Pausa Nov/2026",      "Controles out/2025 suspensos até 10/nov/2026. 7 HREEs ativos. Summit Trump-Xi = catalisador.",               "MP, REMX — upside assimétrico"),
    ("Ucrânia",         "Reconstrução",        "5 rodadas sem avanço. Zelensky 90% fechado. EUA opera como investidor; Europa paga com empresas americanas.", "LMT, RTX, NOC, AECOM"),
    ("Europa / OTAN",   "Rearmamento",         "Alemanha orçamento histórico. Meta 5% PIB. Catalisadores independentes de paz.",                             "LMT, RTX, NOC"),
    ("Tarifas / SCOTUS","Section 232",         "SCOTUS derrubou IEEPA (fev/2026). Section 122 10%→15%. Section 232 constitucional: aço, alumínio, cobre.",   "FCX, CLF, NUE, HG=F"),
    ("Ouro estrutural", "BCs dedolarizando",   "Correlação ouro-dólar quebrou. BCs emergentes 900t/ano. Não é apenas hedge inflacionário — é dedolarização.", "GLD, GC=F, SLV, SI=F"),
]


def style_header_row(ws, row, cols, bg=BG_DARK, fg=AMBER, bold=True, size=10):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill(bg)
        cell.font = font(fg, bold, size)
        cell.alignment = center()
        cell.border = border()


def set_col_widths(ws, widths: list):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ════════════════════════════════════════════════════════════════════════════
# ABA 1 — DASHBOARD
# ════════════════════════════════════════════════════════════════════════════
def build_dashboard(wb: Workbook):
    ws = wb.active
    ws.title = "DASHBOARD"
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = AMBER

    today = datetime.now().strftime("%d/%m/%Y %H:%M")

    # Title block
    ws.merge_cells("A1:H1")
    ws["A1"] = "TRUMP INTEL — MACRO WEEKLY REPORT"
    ws["A1"].fill  = fill(AMBER)
    ws["A1"].font  = Font(color=BG_DARK, bold=True, size=16, name="Arial")
    ws["A1"].alignment = center()

    ws.merge_cells("A2:H2")
    ws["A2"] = f"Gerado em: {today}  •  Edição semanal"
    ws["A2"].fill = fill(BG_CARD)
    ws["A2"].font = font(MUTED, size=9)
    ws["A2"].alignment = center()

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 16

    # ── Macro snapshot ──
    ws.merge_cells("A4:H4")
    ws["A4"] = "COTAÇÕES MACRO"
    ws["A4"].fill = fill(BG_DARK)
    ws["A4"].font = font(AMBER, True, 10)
    ws["A4"].alignment = left()
    ws.row_dimensions[4].height = 20

    macro_headers = ["ATIVO", "SÍMBOLO", "PREÇO", "VAR %", "SETOR", "CONVICÇÃO", "TARGET", "UPSIDE"]
    for ci, h in enumerate(macro_headers, 1):
        c = ws.cell(5, ci, h)
        c.fill = fill("1A1A2A")
        c.font = font(MUTED, True, 9)
        c.alignment = center()
        c.border = border()

    row = 6
    top_items = [item for item in WATCHLIST if item[3] in ("FORTE COMPRA", "COMPRA")][:8]
    for item in top_items:
        sym, finnhub_sym, name, rating, conv, target, cat_label, sector, thesis = item
        q = get_quote(finnhub_sym)
        price = q.get("c") or 0
        chg   = q.get("dp") or 0
        upside = ((target - price) / price * 100) if price > 0 else 0

        rating_color = GREEN if "FORTE" in rating else AMBER if rating == "COMPRA" else MUTED
        chg_color    = GREEN if chg >= 0 else RED
        upside_color = GREEN if upside > 0 else RED

        data = [name, sym,
                f"${price:,.2f}" if price else "—",
                f"{chg:+.2f}%" if chg else "—",
                sector, f"{conv}%",
                f"${target:,.0f}", f"+{upside:.1f}%" if upside > 0 else f"{upside:.1f}%"]

        for ci, val in enumerate(data, 1):
            c = ws.cell(row, ci, val)
            c.fill = fill(BG_CARD)
            c.alignment = center()
            c.border = border()
            if ci == 4:   c.font = font(chg_color,    False, 9)
            elif ci == 6: c.font = font(rating_color, True,  9)
            elif ci == 8: c.font = font(upside_color, True,  9)
            else:         c.font = font(WHITE, False, 9)
        row += 1

    # ── Top conviction ──
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "TOP CONVICÇÃO DESTA SEMANA"
    ws[f"A{row}"].fill = fill(BG_DARK)
    ws[f"A{row}"].font = font(AMBER, True, 10)
    ws[f"A{row}"].alignment = left()
    ws.row_dimensions[row].height = 20
    row += 1

    for item in WATCHLIST[:4]:
        sym, _, name, rating, conv, target, cat_label, sector, thesis = item
        ws.merge_cells(f"A{row}:H{row}")
        ws[f"A{row}"] = f"  {sym} — {name} ({conv}% convicção) → Target: ${target:,}"
        ws[f"A{row}"].fill = fill("0D1A0D")
        ws[f"A{row}"].font = font(GREEN, False, 9)
        ws[f"A{row}"].alignment = left()
        ws[f"A{row}"].border = border()
        row += 1

    # ── Framework reminder ──
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "TRUMP MINDSET: Tarifas = alavanca, não o fim  •  Deal bilateral assimétrico  •  Doutrina Monroe 2.0  •  Hard assets como hedge estrutural"
    ws[f"A{row}"].fill = fill("1A0D00")
    ws[f"A{row}"].font = font(AMBER, False, 9)
    ws[f"A{row}"].alignment = left()
    ws[f"A{row}"].border = border()

    set_col_widths(ws, [22, 12, 12, 10, 14, 12, 12, 10])


# ════════════════════════════════════════════════════════════════════════════
# ABA 2 — WATCHLIST
# ════════════════════════════════════════════════════════════════════════════
def build_watchlist(wb: Workbook):
    ws = wb.create_sheet("WATCHLIST")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = GREEN

    ws.merge_cells("A1:J1")
    ws["A1"] = "WATCHLIST COMPLETA — TRUMP INTEL"
    ws["A1"].fill = fill(BG_DARK)
    ws["A1"].font = font(AMBER, True, 13)
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 28

    headers = ["SÍMBOLO", "NOME", "RATING", "CONVICÇÃO", "PREÇO", "VAR %", "TARGET", "UPSIDE", "CATALISADOR", "SETOR"]
    style_header_row(ws, 2, len(headers))
    for ci, h in enumerate(headers, 1):
        ws.cell(2, ci, h)

    for row_i, item in enumerate(WATCHLIST, 3):
        sym, finnhub_sym, name, rating, conv, target, cat_label, sector, thesis = item
        q = get_quote(finnhub_sym)
        price  = q.get("c") or 0
        chg    = q.get("dp") or 0
        upside = ((target - price) / price * 100) if price > 0 else 0

        bg = "0D1A0D" if "FORTE" in rating else "0D0D1A" if rating == "COMPRA" else BG_CARD
        rating_color = GREEN if "FORTE" in rating else AMBER if rating == "COMPRA" else MUTED
        chg_color    = GREEN if chg >= 0 else RED

        row_data = [
            sym, name, rating, f"{conv}%",
            f"${price:,.2f}" if price else "—",
            f"{chg:+.2f}%" if chg else "—",
            f"${target:,.0f}",
            f"+{upside:.1f}%" if upside > 0 else f"{upside:.1f}%",
            cat_label, sector
        ]
        for ci, val in enumerate(row_data, 1):
            c = ws.cell(row_i, ci, val)
            c.fill = fill(bg)
            c.alignment = center()
            c.border = border()
            if ci == 3:   c.font = font(rating_color, True, 9)
            elif ci == 6: c.font = font(chg_color, False, 9)
            elif ci == 8: c.font = font(GREEN if upside > 10 else AMBER, True, 9)
            else:         c.font = font(WHITE, False, 9)

        ws.row_dimensions[row_i].height = 16

    set_col_widths(ws, [10, 22, 14, 12, 12, 10, 12, 10, 14, 12])

    # Thesis block below
    row = len(WATCHLIST) + 4
    ws.merge_cells(f"A{row}:J{row}")
    ws[f"A{row}"] = "TESES RESUMIDAS"
    ws[f"A{row}"].fill = fill(BG_DARK)
    ws[f"A{row}"].font = font(AMBER, True, 10)
    ws[f"A{row}"].alignment = left()

    for item in WATCHLIST:
        row += 1
        sym, _, name, rating, conv, target, _, sector, thesis = item
        ws.merge_cells(f"A{row}:J{row}")
        ws[f"A{row}"] = f"  {sym} ({conv}%): {thesis}"
        ws[f"A{row}"].fill = fill(BG_CARD)
        ws[f"A{row}"].font = font(MUTED, False, 9)
        ws[f"A{row}"].alignment = left()
        ws[f"A{row}"].border = border()
        ws.row_dimensions[row].height = 15


# ════════════════════════════════════════════════════════════════════════════
# ABA 3 — TESES GEOPOLÍTICAS
# ════════════════════════════════════════════════════════════════════════════
def build_teses(wb: Workbook):
    ws = wb.create_sheet("TESES")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = RED

    ws.merge_cells("A1:E1")
    ws["A1"] = "TESES GEOPOLÍTICAS — DOUTRINA TRUMP"
    ws["A1"].fill = fill(BG_DARK)
    ws["A1"].font = font(AMBER, True, 13)
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 28

    headers = ["TEATRO", "TESE CENTRAL", "ANÁLISE", "AÇÃO / PLAY", "STATUS"]
    style_header_row(ws, 2, len(headers))
    for ci, h in enumerate(headers, 1):
        ws.cell(2, ci, h)

    status_map = {
        "Venezuela":       ("ATIVO", RED),
        "China — Raros":   ("SUMMIT 22d", AMBER),
        "Ucrânia":         ("NEGOC.", AMBER),
        "Europa / OTAN":   ("WATCH", BLUE),
        "Tarifas / SCOTUS":("SECTION 232", BLUE),
        "Ouro estrutural": ("ESTRUTURAL", GREEN),
    }

    for row_i, (teatro, tese, analise, play) in enumerate(TESES, 3):
        status_label, status_color = status_map.get(teatro, ("WATCH", MUTED))
        row_data = [teatro, tese, analise, play, status_label]
        for ci, val in enumerate(row_data, 1):
            c = ws.cell(row_i, ci, val)
            c.fill = fill("0D0D1A" if row_i % 2 == 0 else BG_CARD)
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            c.border = border()
            if ci == 1: c.font = font(AMBER, True, 9)
            elif ci == 5: c.font = font(status_color, True, 9)
            else: c.font = font(WHITE, False, 9)
        ws.row_dimensions[row_i].height = 45

    set_col_widths(ws, [18, 22, 50, 30, 14])


# ════════════════════════════════════════════════════════════════════════════
# ABA 4 — GATILHOS / CALENDÁRIO
# ════════════════════════════════════════════════════════════════════════════
def build_gatilhos(wb: Workbook):
    ws = wb.create_sheet("GATILHOS")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = RED

    ws.merge_cells("A1:E1")
    ws["A1"] = "CALENDÁRIO DE GATILHOS 2026"
    ws["A1"].fill = fill(BG_DARK)
    ws["A1"].font = font(AMBER, True, 13)
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 28

    headers = ["DATA", "EVENTO", "IMPACTO", "DETALHE", "ATIVOS RELACIONADOS"]
    style_header_row(ws, 2, len(headers))
    for ci, h in enumerate(headers, 1):
        ws.cell(2, ci, h)

    impact_colors = {"CRÍTICO": RED, "ALTO": AMBER, "MÉDIO": BLUE}

    for row_i, (data, evento, impacto, detalhe, ativos) in enumerate(CATALYSTS, 3):
        color = impact_colors.get(impacto, MUTED)
        row_data = [data, evento, impacto, detalhe, ativos]
        for ci, val in enumerate(row_data, 1):
            c = ws.cell(row_i, ci, val)
            c.fill = fill("1A0000" if impacto == "CRÍTICO" else "1A1000" if impacto == "ALTO" else BG_CARD)
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            c.border = border()
            if ci == 3: c.font = font(color, True, 10)
            elif ci == 2: c.font = font(WHITE, True, 9)
            else: c.font = font(WHITE, False, 9)
        ws.row_dimensions[row_i].height = 35

    set_col_widths(ws, [18, 28, 12, 50, 24])


# ════════════════════════════════════════════════════════════════════════════
# ABA 5 — HISTÓRICO (lê briefings salvos)
# ════════════════════════════════════════════════════════════════════════════
def build_historico(wb: Workbook):
    ws = wb.create_sheet("HISTÓRICO")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = BLUE

    ws.merge_cells("A1:C1")
    ws["A1"] = "HISTÓRICO DE BRIEFINGS"
    ws["A1"].fill = fill(BG_DARK)
    ws["A1"].font = font(AMBER, True, 13)
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 28

    headers = ["DATA", "GERADO EM", "BRIEFING"]
    style_header_row(ws, 2, 3)
    for ci, h in enumerate(headers, 1):
        ws.cell(2, ci, h)

    # Try to load current briefing
    row = 3
    briefing_path = "data/briefing.json"
    if os.path.exists(briefing_path):
        try:
            with open(briefing_path, "r", encoding="utf-8") as f:
                b = json.load(f)
            for ci, val in enumerate([b.get("date","—"), b.get("generated_at","—")[:16], b.get("text","—")], 1):
                c = ws.cell(row, ci, val)
                c.fill = fill(BG_CARD)
                c.font = font(WHITE, False, 9)
                c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                c.border = border()
            ws.row_dimensions[row].height = 80
            row += 1
        except Exception:
            pass

    # Placeholder rows
    for _ in range(4):
        for ci in range(1, 4):
            c = ws.cell(row, ci, "—" if ci < 3 else "")
            c.fill = fill(BG_CARD)
            c.font = font(MUTED, False, 9)
            c.border = border()
        ws.row_dimensions[row].height = 20
        row += 1

    set_col_widths(ws, [14, 20, 90])
    ws.column_dimensions["C"].width = 90


# ════════════════════════════════════════════════════════════════════════════
# MAIN
# ════════════════════════════════════════════════════════════════════════════
def main():
    print("Gerando relatório semanal...")
    wb = Workbook()

    build_dashboard(wb)
    build_watchlist(wb)
    build_teses(wb)
    build_gatilhos(wb)
    build_historico(wb)

    os.makedirs("data", exist_ok=True)
    dated_name  = f"data/report_{datetime.now().strftime('%Y%m%d')}.xlsx"
    latest_name = "data/report_latest.xlsx"

    wb.save(dated_name)
    wb.save(latest_name)
    print(f"✓ Salvo: {dated_name}")
    print(f"✓ Salvo: {latest_name}")


if __name__ == "__main__":
    main()
